
"""file=open("sample.txt","w")
file.write("This is a sample file created using python.")
file.close()
file=open("sample.txt","r")
content=file.read()
print(content)
file.close()
with open("sample.txt","r") as file:
    content=file.read()
print(content)

try:
    with open("missing.txt","r") as file:
        print(file.read())
except FileNotFoundError:
    print("The file 'missing.txt' was not found.")
"""
"""
import csv
with open("D:\\DS_AI_Internship\\src\\Day7\\data.csv","r")as file:
    reader=csv.reader(file)
    for row in reader:
        print(row)
"""

from openpyxl import load_workbook
file_path_excel = "D:\\DS_AI_Internship\\src\\Day7\\data.xlsx"
workbook = load_workbook(file_path_excel)
sheet = workbook.active  # first sheet
for row in sheet.iter_rows(values_only=True):
        print(row)
